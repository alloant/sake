#!/bin/python
import logging
from attrdict import AttrDict

from openpyxl import load_workbook

from libsynomail.nas import download_path

class Register(AttrDict):
    def __init__(self,flow,path_archive):
        file = download_path(f"{path_archive}/Mail {flow} Registry.osheet")
        if not file:
            logging.error("Cannot read Mail out Registry")
            return None

        self.wb = load_workbook(file)
        self.cg = list(self.wb[f'cg {flow} (1-249)'].iter_rows(values_only=True))
        self.asr = list(self.wb[f'asr {flow} (250-999)'].iter_rows(values_only=True))
        self.ctr = list(self.wb[f'ctr {flow} (from 1000 to 1999)'].iter_rows(values_only=True))
        self.r = list(self.wb[f'r {flow} (2000 onwards)'].iter_rows(values_only=True))

    def get_type(self,no):
        if no < 250:
            tp = 'cg'
        elif no < 1000:
            tp = 'asr'
        elif no < 2000:
            tp = 'ctr'
        else:
            tp = 'r'

        return tp

    def scrap_destination(self,no):
        found = False
        for reg in self[self.get_type(no)]:
            if reg[0] == no:
                found = True
                break

        if self.get_type(no) in ['ctr','r']:
            if not found: return '','',''
            return reg[2],reg[4],reg[3]
        else:
            if not found: return self.get_type(no),'',''
            
            if self.get_type(no) == 'cg':
                return self.get_type(no),reg[4],reg[3]
            else:
                return self.get_type(no),reg[3],reg[2]


 
