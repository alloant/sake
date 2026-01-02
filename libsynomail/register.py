#!/bin/python
from datetime import datetime
from tempfile import NamedTemporaryFile
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

import webbrowser

from libsynomail.nas import download_path, upload_path, convert_office, files_path
from libsynomail.classes import Note, File

FONT = Font(name= 'Arial',
                size=12,
                bold=False,
                italic=False,
                strike=False,
                underline='none'
                #color='4472C4'
                )

FONT_BOLD = Font(name= 'Arial',
                size=12,
                bold=True,
                italic=False,
                strike=False,
                underline='none'
                #color='4472C4'
                )


def write_register(path,notes,browser = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    ws.append(['Register','Type','Source','No','Year','Ref','Date','Content','Dept','of_annex','Comments','Archived','Sent to'])
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = FONT_BOLD
    
    ws_data = wb.create_sheet(title="Data")
    ws_data.append(['Key','Folder_path','Folder_id','Permanent_link'])
 
    ws_files = wb.create_sheet(title="Files")
    ws_files.append(['Key','Name','Type','Display_data','File_id','Permanent_link','Original_path','Original_id'])

    
    for key,note in notes.items():
        ws.append(note.exportExcel())
        ws[ws.max_row][6].number_format = 'dd/mm/yyyy;@'
       
        ws_data.append([note.key,note.folder_path,note.folder_id,note.permanent_link])

        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center')
            cell.font = FONT

        for file in note.files:
            ws_files.append([key] + file.exportExcel())

    column_widths = [10,10,10,10,12,12,15,50,20,12,50,10,20]
    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = column_width
    
    date = datetime.today().strftime('%Y-%m-%d-%HH-%Mm')
    upload_register(wb,f"register-{date}.xlsx",f"{path}",browser)

def upload_register(wb,name,dest,browser = None):
    try:
        file = NamedTemporaryFile()
        wb.save(file)
        file.seek(0)
        file.name = name

        ret_upload = upload_path(file,dest)
        uploaded = True if ret_upload else False
    except Exception as err:
        logging.error(err)
        logging.error("Cannot upload register")
        wb.save(file.name)
        uploaded = False

    if uploaded:
        try:
            rst = convert_office(ret_upload['data']['display_path'],delete=False)
            if rst:
                if browser == None:
                    webbrowser.open(f"https://nas.prome.sg:5001/oo/r/{rst['permanent_link']}")
                else:
                    mybrowser = webbrowser.get(browser)
                    mybrowser.open(f"https://nas.prome.sg:5001/oo/r/{rst['permanent_link']}")
        except Exception as err:
            logging.error(err)
            logging.warning("Cannot convert register to Synology Office")
    else:
        logging.error("Cannot upload register")
        wb.save(file.name)


def read_register(path_despacho,flow = 'in'):
    files_in_outbox = files_path(f"{path_despacho}")
    
    if not files_in_outbox: return False

    files_in_outbox.sort(reverse = True,key = lambda file: file['name'])
    notes = ''
    for file in files_in_outbox:
        if file['name'][:9] in ["register-"] and 'osheet' in file['name']:
            notes = register_to_notes(file['display_path'],flow)
            break

    return notes

def join_registers(path,flow = 'in',memory_notes = {},browser = None):
    files_in_outbox = files_path(f"{path}")
    
    if not files_in_outbox: return False

    files_in_outbox.sort(key = lambda file: file['name'])
    notes = memory_notes
    for file in files_in_outbox:
        if file['name'][:9] in ["register-"] and 'osheet' in file['name']:
            nts = register_to_notes(file['display_path'],flow)

            notes |= nts

    write_register(path,notes,browser)
            
def register_to_notes(register,flow = 'in'):
    dow_register = download_path(register)
    if dow_register:
        wb = load_workbook(dow_register)
    else:
        return False

    notes_data = list(wb['Notes'].iter_rows(values_only=True))[1:]
    data = list(wb['Data'].iter_rows(values_only=True))[1:]
    files = list(wb['Files'].iter_rows(values_only=True))[1:]
    
    notes = {}
    
    for i,row in enumerate(notes_data):
        if row[0] == None: continue
        no = row[3].replace(' ','').split('","')[1][:-2]
        note = Note(row[0],row[1],row[2],no,flow=flow)
        note.year = row[4] if row[4] else ''
        note.ref = row[5] if row[5] else ''
        note.date = row[6] if row[6] else '' 
        note.content = row[7] if row[7] else ''
        note.dept = row[8] if row[8] else '' 
        note.comments = row[10] if row[10] else ''
        note.archived = row[11] if row[11] else ''
        note.sent_to = row[12] if row[12] else ''

        note.folder_path = data[i][1]
        note.folder_id = data[i][2]
        note.permanent_link = data[i][3]

        for file in files:
            if file[0] == note.key:
                note.addFile(File({'name':file[1],'type':file[2],'display_path':file[3],'file_id':file[4],'permanent_link':file[5]},file[6],file[7]))

        notes[note.key] = note
   
    return notes

